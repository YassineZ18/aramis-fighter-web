#!/usr/bin/env python3
"""
Générateur de template Excel pour l'analyse d'escrime Aramis Fighter
Alternative Python utilisant openpyxl pour créer un template Excel avancé
"""

import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import os

class AramisExcelTemplate:
    def __init__(self):
        self.workbook = openpyxl.Workbook()
        self.setup_styles()
    
    def setup_styles(self):
        """Définit les styles réutilisables"""
        self.header_font = Font(bold=True, color="FFFFFF")
        self.title_font = Font(size=16, bold=True, color="4472C4")
        self.subtitle_font = Font(size=12, bold=True, color="2E86AB")
        
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.analysis_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        self.actions_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def create_data_sheet(self):
        """Feuille 1: Données brutes importées"""
        ws = self.workbook.active
        ws.title = "Données Brutes"
        
        # En-têtes
        headers = [
            'Date', 'Heure', 'Escrimeur_1', 'Escrimeur_2', 'Score_1', 'Score_2',
            'Touche_Num', 'Escrimeur_Touchant', 'Action_Code', 'Action_Nom',
            'Zone_Touchee', 'Validite', 'Duree_Action', 'Efficacite'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Ajustement des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Figer la première ligne
        ws.freeze_panes = 'A2'
        
        return ws
    
    def create_analysis_sheet(self):
        """Feuille 2: Analyse par escrimeur"""
        ws = self.workbook.create_sheet("Analyse Escrimeur")
        
        # Titre principal
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'ANALYSE DÉTAILLÉE PAR ESCRIMEUR'
        title_cell.font = self.title_font
        title_cell.alignment = self.center_alignment
        
        # En-têtes d'analyse
        headers = [
            'Escrimeur', 'Matchs_Joués', 'Victoires', 'Défaites', 'Ratio_Victoires',
            'Touches_Données', 'Touches_Reçues', 'Efficacité_Globale'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.analysis_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Exemple avec formules
        example_data = [
            'Exemple_Escrimeur',
            '=COUNTIFS(\'Données Brutes\'.C:C,A3)+COUNTIFS(\'Données Brutes\'.D:D,A3)',
            '=SUMPRODUCT((\'Données Brutes\'.C:C=A3)*(\'Données Brutes\'.E:E>\'Données Brutes\'.F:F)+(\'Données Brutes\'.D:D=A3)*(\'Données Brutes\'.F:F>\'Données Brutes\'.E:E))',
            '=B3-C3',
            '=IF(B3>0,C3/B3,0)',
            '=COUNTIFS(\'Données Brutes\'.H:H,A3)',
            '=COUNTIFS(\'Données Brutes\'.C:C,A3,\'Données Brutes\'.H:H,"<>"&A3)+COUNTIFS(\'Données Brutes\'.D:D,A3,\'Données Brutes\'.H:H,"<>"&A3)',
            '=IF(F3+G3>0,F3/(F3+G3),0)'
        ]
        
        for col, value in enumerate(example_data, 1):
            cell = ws.cell(row=3, column=col, value=value)
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            cell.font = Font(italic=True)
        
        # Ajustement des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        ws.freeze_panes = 'A3'
        return ws
    
    def create_actions_sheet(self):
        """Feuille 3: Analyse des actions"""
        ws = self.workbook.create_sheet("Analyse Actions")
        
        # Titre
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = 'ANALYSE DES ACTIONS D\'ESCRIME'
        title_cell.font = Font(size=16, bold=True, color="E74C3C")
        title_cell.alignment = self.center_alignment
        
        # En-têtes
        headers = ['Action_Code', 'Action_Nom', 'Fréquence', 'Efficacité', 'Zone_Préférée', 'Recommandation']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.actions_fill
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
        
        # Données d'exemple
        action_data = [
            ['ATT_D', 'Attaque Directe', '=COUNTIFS(\'Données Brutes\'.I:I,"ATT_D")', '=COUNTIFS(\'Données Brutes\'.I:I,"ATT_D",\'Données Brutes\'.L:L,"Valide")/C3', 'Torse', 'Action efficace - Maintenir'],
            ['RIP_D', 'Riposte Directe', '=COUNTIFS(\'Données Brutes\'.I:I,"RIP_D")', '=COUNTIFS(\'Données Brutes\'.I:I,"RIP_D",\'Données Brutes\'.L:L,"Valide")/C4', 'Bras', 'Améliorer précision'],
            ['PAR_RIP', 'Parade-Riposte', '=COUNTIFS(\'Données Brutes\'.I:I,"PAR_RIP")', '=COUNTIFS(\'Données Brutes\'.I:I,"PAR_RIP",\'Données Brutes\'.L:L,"Valide")/C5', 'Torse', 'Excellente défense'],
            ['CTR_ATT', 'Contre-Attaque', '=COUNTIFS(\'Données Brutes\'.I:I,"CTR_ATT")', '=COUNTIFS(\'Données Brutes\'.I:I,"CTR_ATT",\'Données Brutes\'.L:L,"Valide")/C6', 'Bras', 'Timing à améliorer'],
            ['ATT_C', 'Attaque Composée', '=COUNTIFS(\'Données Brutes\'.I:I,"ATT_C")', '=COUNTIFS(\'Données Brutes\'.I:I,"ATT_C",\'Données Brutes\'.L:L,"Valide")/C7', 'Torse', 'Complexité maîtrisée']
        ]
        
        for row_idx, row_data in enumerate(action_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Mise en forme conditionnelle pour l'efficacité
        rule = ColorScaleRule(
            start_type='min', start_value=0, start_color='FF6B6B',
            mid_type='percentile', mid_value=50, mid_color='FFEB3B',
            end_type='max', end_value=1, end_color='4ECDC4'
        )
        ws.conditional_formatting.add('D3:D7', rule)
        
        # Ajustement des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        ws.freeze_panes = 'A3'
        return ws
    
    def create_charts_sheet(self):
        """Feuille 4: Graphiques et visualisations"""
        ws = self.workbook.create_sheet("Graphiques")
        
        # Masquer les lignes de grille
        ws.sheet_view.showGridLines = False
        
        # Titre
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = 'VISUALISATIONS ET GRAPHIQUES'
        title_cell.font = Font(size=18, bold=True, color="9B59B6")
        title_cell.alignment = self.center_alignment
        
        # Zone graphique 1
        ws.merge_cells('A3:D15')
        chart1_cell = ws['A3']
        chart1_cell.value = 'GRAPHIQUE 1: Répartition des Actions\n\n(Graphique en secteurs)\n\nDonnées: Feuille "Analyse Actions"\nColonnes: Action_Nom, Fréquence'
        chart1_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        chart1_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        chart1_cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        # Zone graphique 2
        ws.merge_cells('F3:H15')
        chart2_cell = ws['F3']
        chart2_cell.value = 'GRAPHIQUE 2: Efficacité par Action\n\n(Graphique en barres)\n\nDonnées: Feuille "Analyse Actions"\nColonnes: Action_Nom, Efficacité'
        chart2_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        chart2_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        chart2_cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        # Zone graphique 3
        ws.merge_cells('A17:H25')
        chart3_cell = ws['A17']
        chart3_cell.value = 'GRAPHIQUE 3: Évolution des Performances\n\n(Graphique linéaire)\n\nDonnées: Feuille "Données Brutes"\nAxe X: Date, Axe Y: Efficacité par match'
        chart3_cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        chart3_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        chart3_cell.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        
        return ws
    
    def create_macros_sheet(self):
        """Feuille 5: Instructions pour macros"""
        ws = self.workbook.create_sheet("Macros & Auto")
        ws.sheet_view.showGridLines = False
        
        instructions = [
            ['INSTRUCTIONS POUR LES MACROS EXCEL'],
            [''],
            ['1. MACRO IMPORT_CSV:'],
            ['   - Ouvre le sélecteur de fichier CSV'],
            ['   - Importe automatiquement dans "Données Brutes"'],
            ['   - Met à jour tous les calculs'],
            [''],
            ['2. MACRO REFRESH_ANALYSIS:'],
            ['   - Recalcule toutes les formules'],
            ['   - Met à jour les graphiques'],
            ['   - Applique la mise en forme conditionnelle'],
            [''],
            ['3. MACRO EXPORT_REPORT:'],
            ['   - Génère un rapport PDF'],
            ['   - Exporte les graphiques'],
            ['   - Sauvegarde avec timestamp'],
            [''],
            ['CODE VBA À AJOUTER:'],
            [''],
            ['Sub Import_CSV()'],
            ['    Dim filePath As String'],
            ['    filePath = Application.GetOpenFilename("CSV Files (*.csv), *.csv")'],
            ['    If filePath <> "False" Then'],
            ['        Workbooks.OpenText filePath, DataType:=xlDelimited, Comma:=True'],
            ['        \' Copier données vers feuille "Données Brutes"'],
            ['        \' Code d\'import personnalisé ici'],
            ['    End If'],
            ['End Sub'],
            [''],
            ['Sub Refresh_Analysis()'],
            ['    Application.CalculateFullRebuild'],
            ['    \' Mise à jour des graphiques'],
            ['    \' Code de rafraîchissement ici'],
            ['End Sub']
        ]
        
        for row_idx, instruction in enumerate(instructions, 1):
            cell = ws.cell(row=row_idx, column=1, value=instruction[0] if instruction else '')
            
            if row_idx == 1:
                cell.font = Font(size=16, bold=True, color="2E86AB")
                cell.alignment = self.center_alignment
                ws.merge_cells(f'A{row_idx}:F{row_idx}')
            elif instruction and instruction[0] and ('MACRO' in instruction[0] or 'CODE VBA' in instruction[0]):
                cell.font = self.subtitle_font
            elif instruction and instruction[0] and ('Sub ' in instruction[0] or 'End Sub' in instruction[0] or instruction[0].startswith('    ')):
                cell.font = Font(name='Consolas', size=10, color="666666")
                cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        ws.column_dimensions['A'].width = 80
        return ws
    
    def generate_template(self):
        """Génère le template Excel complet"""
        print("🔧 Génération du template Excel Aramis Fighter...")
        
        # Supprimer la feuille par défaut et créer toutes les feuilles
        self.workbook.remove(self.workbook.active)
        
        self.create_data_sheet()
        self.create_analysis_sheet()
        self.create_actions_sheet()
        self.create_charts_sheet()
        self.create_macros_sheet()
        
        # Sauvegarder le fichier
        template_path = os.path.join(os.path.dirname(__file__), 'Aramis_Fighter_Analysis_Template.xlsx')
        
        try:
            self.workbook.save(template_path)
            print(f"✅ Template Excel généré avec succès: {template_path}")
            return template_path
        except Exception as error:
            print(f"❌ Erreur lors de la génération du template: {error}")
            raise error

if __name__ == "__main__":
    generator = AramisExcelTemplate()
    try:
        path = generator.generate_template()
        print(f"🎉 Template prêt à utiliser: {path}")
    except Exception as error:
        print(f"💥 Échec de la génération: {error}")
