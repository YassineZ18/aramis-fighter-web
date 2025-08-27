#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Analyseur de fichier Excel pour extraire la structure des graphiques
Analyse: Analyse assauts Vichy 2022 EPEE.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl.chart import *
import json
import sys

def analyze_excel_file(file_path):
    """Analyse la structure du fichier Excel et extrait les informations sur les graphiques"""
    
    results = {
        "sheets": [],
        "charts": [],
        "data_structure": {},
        "chart_types": [],
        "recommendations": []
    }
    
    try:
        # Charger le workbook avec openpyxl pour analyser les graphiques
        wb = openpyxl.load_workbook(file_path)
        
        print("📊 ANALYSE DU FICHIER EXCEL VICHY 2022")
        print("=" * 50)
        
        # Analyser chaque feuille
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "charts": [],
                "data_ranges": [],
                "max_row": sheet.max_row,
                "max_col": sheet.max_column
            }
            
            print(f"\n📋 Feuille: {sheet_name}")
            print(f"   Dimensions: {sheet.max_row} lignes x {sheet.max_column} colonnes")
            
            # Analyser les graphiques dans cette feuille
            for chart in sheet._charts:
                chart_info = {
                    "type": type(chart).__name__,
                    "title": getattr(chart, 'title', None),
                    "anchor": str(chart.anchor),
                    "series_count": len(chart.series) if hasattr(chart, 'series') else 0
                }
                
                print(f"   📈 Graphique: {chart_info['type']}")
                if chart_info['title']:
                    print(f"      Titre: {chart_info['title']}")
                print(f"      Séries: {chart_info['series_count']}")
                
                sheet_info["charts"].append(chart_info)
                
                # Ajouter le type de graphique à la liste globale
                if chart_info['type'] not in results["chart_types"]:
                    results["chart_types"].append(chart_info['type'])
            
            results["sheets"].append(sheet_info)
        
        # Analyser les données avec pandas
        print(f"\n📊 ANALYSE DES DONNÉES")
        print("=" * 30)
        
        try:
            # Lire toutes les feuilles
            excel_data = pd.read_excel(file_path, sheet_name=None, header=None)
            
            for sheet_name, df in excel_data.items():
                print(f"\n📋 Données - {sheet_name}:")
                print(f"   Shape: {df.shape}")
                
                # Chercher des patterns de données d'escrime
                escrime_keywords = ['AS', 'AC', 'AF', 'PR', 'CA', 'Zone', 'Efficacité', 'Touches']
                
                for keyword in escrime_keywords:
                    matches = df.astype(str).apply(lambda x: x.str.contains(keyword, case=False, na=False)).any()
                    if matches.any():
                        print(f"   ✓ Trouvé: {keyword}")
                
                # Sauvegarder un échantillon des premières lignes
                sample = df.head(10).fillna('').astype(str)
                results["data_structure"][sheet_name] = sample.to_dict()
        
        except Exception as e:
            print(f"   ⚠️ Erreur lecture pandas: {e}")
        
        # Générer des recommandations
        print(f"\n🎯 RECOMMANDATIONS POUR ARAMIS FIGHTER")
        print("=" * 40)
        
        chart_recommendations = {
            "BarChart": "Graphiques en barres pour efficacité par action",
            "PieChart": "Graphiques circulaires pour répartition des actions", 
            "LineChart": "Courbes d'évolution des performances",
            "ScatterChart": "Nuages de points pour corrélations",
            "AreaChart": "Graphiques en aires pour zones de terrain",
            "RadarChart": "Graphiques radar pour profils tactiques"
        }
        
        for chart_type in results["chart_types"]:
            if chart_type in chart_recommendations:
                recommendation = chart_recommendations[chart_type]
                results["recommendations"].append({
                    "chart_type": chart_type,
                    "description": recommendation
                })
                print(f"   📈 {chart_type}: {recommendation}")
        
        # Sauvegarder l'analyse
        with open('excel_analysis_results.json', 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n✅ Analyse sauvegardée dans: excel_analysis_results.json")
        
        return results
        
    except Exception as e:
        print(f"❌ Erreur lors de l'analyse: {e}")
        return None

if __name__ == "__main__":
    file_path = "Analyse assauts Vichy 2022 EPEE.xlsx"
    analyze_excel_file(file_path)
